#!/usr/bin/env python3
import argparse
import json
import platform
import re
import statistics
import sys
import tempfile
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import requests
from openpyxl import Workbook
from openpyxl.styles import Font

try:
    from playwright.sync_api import sync_playwright
except Exception:  # pragma: no cover
    sync_playwright = None


USER_BY_SCREEN_NAME_FEATURES = {
    "hidden_profile_subscriptions_enabled": True,
    "profile_label_improvements_pcf_label_in_post_enabled": True,
    "responsive_web_profile_redirect_enabled": True,
    "rweb_tipjar_consumption_enabled": True,
    "verified_phone_label_enabled": True,
    "subscriptions_verification_info_is_identity_verified_enabled": True,
    "subscriptions_verification_info_verified_since_enabled": True,
    "highlights_tweets_tab_ui_enabled": True,
    "responsive_web_twitter_article_notes_tab_enabled": True,
    "subscriptions_feature_can_gift_premium": True,
    "creator_subscriptions_tweet_preview_api_enabled": True,
    "responsive_web_graphql_skip_user_profile_image_extensions_enabled": False,
    "responsive_web_graphql_timeline_navigation_enabled": True,
}
USER_BY_SCREEN_NAME_FIELD_TOGGLES = {"withPayments": False, "withAuxiliaryUserLabels": False}
USER_TWEETS_FEATURES = {
    "rweb_video_screen_enabled": False,
    "profile_label_improvements_pcf_label_in_post_enabled": True,
    "responsive_web_profile_redirect_enabled": True,
    "rweb_tipjar_consumption_enabled": True,
    "verified_phone_label_enabled": True,
    "creator_subscriptions_tweet_preview_api_enabled": True,
    "responsive_web_graphql_timeline_navigation_enabled": True,
    "responsive_web_graphql_skip_user_profile_image_extensions_enabled": False,
    "premium_content_api_read_enabled": False,
    "communities_web_enable_tweet_community_results_fetch": True,
    "c9s_tweet_anatomy_moderator_badge_enabled": True,
    "responsive_web_grok_analyze_button_fetch_trends_enabled": False,
    "responsive_web_grok_analyze_post_followups_enabled": True,
    "responsive_web_jetfuel_frame": False,
    "responsive_web_grok_share_attachment_enabled": True,
    "responsive_web_grok_annotations_enabled": True,
    "articles_preview_enabled": True,
    "responsive_web_edit_tweet_api_enabled": True,
    "graphql_is_translatable_rweb_tweet_is_translatable_enabled": True,
    "view_counts_everywhere_api_enabled": True,
    "longform_notetweets_consumption_enabled": True,
    "responsive_web_twitter_article_tweet_consumption_enabled": True,
    "tweet_awards_web_tipping_enabled": False,
    "responsive_web_grok_analysis_button_from_backend": True,
    "creator_subscriptions_quote_tweet_preview_enabled": False,
    "freedom_of_speech_not_reach_fetch_enabled": True,
    "standardized_nudges_misinfo": True,
    "tweet_with_visibility_results_prefer_gql_limited_actions_policy_enabled": True,
    "longform_notetweets_rich_text_read_enabled": True,
    "longform_notetweets_inline_media_enabled": True,
    "responsive_web_grok_image_annotation_enabled": True,
    "responsive_web_grok_imagine_annotation_enabled": True,
    "responsive_web_grok_community_note_auto_translation_is_enabled": False,
    "responsive_web_enhance_cards_enabled": False,
}
USER_TWEETS_FIELD_TOGGLES = {
    "withPayments": False,
    "withAuxiliaryUserLabels": False,
    "withArticleRichContentState": True,
    "withArticlePlainText": False,
    "withGrokAnalyze": False,
    "withDisallowedReplyControls": False,
}
ALLOWED_POST_TYPES = {"article", "post"}
EXCLUDED_REASONS = {"quote", "note", "reply", "retweet", "thread_continuation"}
SUMMARY_HEADERS = [
    "handle",
    "display_name",
    "profile_url",
    "followers",
    "price",
    "source_used",
    "sample_size",
    "average_impression",
    "median_impression",
    "average_engagement",
    "median_engagement",
    "engagement_rate",
    "median_engagement_rate",
    "cost_per_impression",
    "usd_cpm",
    "assessment",
]
DETAIL_HEADERS = [
    "handle",
    "post_id",
    "post_url",
    "created_at",
    "post_type",
    "impressions",
    "likes",
    "retweets",
    "replies",
    "engagement",
    "engagement_rate",
    "selection_status",
    "exclusion_reason",
    "text_preview",
]


@dataclass
class Target:
    handle: str
    profile_url: str
    price: Optional[str] = None


class SkillError(RuntimeError):
    pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["account_list", "discovery"], required=True)
    parser.add_argument("--targets", default="")
    parser.add_argument("--targets-file")
    parser.add_argument("--topic")
    parser.add_argument("--count", type=int, default=10)
    parser.add_argument("--product-context", default="")
    parser.add_argument("--browser-profile")
    parser.add_argument("--headed", action="store_true")
    parser.add_argument("--allow-partial", action="store_true")
    parser.add_argument("--dependency-bootstrap")
    return parser.parse_args()


def compact_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def slugify(value: str) -> str:
    return re.sub(r"[^a-zA-Z0-9_-]+", "-", value.strip()).strip("-").lower() or "run"


def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def write_json(path: Path, payload: Any) -> None:
    ensure_dir(path.parent)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n")


def parse_price(label: Optional[str]) -> Tuple[float, str]:
    label = label or ""
    if "$" in label:
        vals = [float(x) for x in re.findall(r"\$([0-9]+(?:\.[0-9]+)?)", label)]
        if vals:
            return min(vals), "$"
    if "€" in label:
        vals = [float(x) for x in re.findall(r"€([0-9]+(?:\.[0-9]+)?)", label)]
        if vals:
            return min(vals), "€"
    vals = [float(x) for x in re.findall(r"([0-9]+(?:\.[0-9]+)?)\s*RMB", label, flags=re.I)]
    if vals:
        return min(vals), "RMB "
    return 0.0, ""


def extract_price_points(label: Optional[str]) -> List[Tuple[float, str]]:
    label = label or ""
    points: List[Tuple[float, str]] = []
    for value in re.findall(r"\$([0-9]+(?:\.[0-9]+)?)", label):
        points.append((float(value), "USD"))
    for value in re.findall(r"€([0-9]+(?:\.[0-9]+)?)", label):
        points.append((float(value), "EUR"))
    for value in re.findall(r"£([0-9]+(?:\.[0-9]+)?)", label):
        points.append((float(value), "GBP"))
    for value in re.findall(r"([0-9]+(?:\.[0-9]+)?)\s*(?:USD|USDT|dollars?)", label, flags=re.I):
        points.append((float(value), "USD"))
    for value in re.findall(r"([0-9]+(?:\.[0-9]+)?)\s*(?:EUR)", label, flags=re.I):
        points.append((float(value), "EUR"))
    for value in re.findall(r"([0-9]+(?:\.[0-9]+)?)\s*(?:GBP)", label, flags=re.I):
        points.append((float(value), "GBP"))
    for value in re.findall(r"([0-9]+(?:\.[0-9]+)?)\s*(?:RMB|CNY|CNH)", label, flags=re.I):
        points.append((float(value), "CNY"))
    return points


def fetch_google_finance_rate(page: Any, pair: str) -> float:
    page.goto(f"https://www.google.com/finance/quote/{pair}?hl=en", wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(1200)
    try:
        page.get_by_role("button", name="Reject all").first.click(timeout=2500)
        page.wait_for_timeout(1200)
    except Exception:
        pass
    text = page.locator("body").inner_text(timeout=10000)
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    marker = {
        "EUR-USD": "EUR / USD • CURRENCY",
        "GBP-USD": "GBP / USD • CURRENCY",
        "USD-CNY": "USD / CNY • CURRENCY",
    }[pair]
    if marker not in lines:
        raise SkillError(f"Could not find Google Finance quote marker for {pair}")
    idx = lines.index(marker)
    for offset in range(idx + 1, min(idx + 8, len(lines))):
        candidate = lines[offset].replace(",", "")
        if re.fullmatch(r"[0-9]+(?:\.[0-9]+)?", candidate):
            return float(candidate)
    raise SkillError(f"Could not parse Google Finance quote for {pair}")


def fetch_fx_rates_to_usd() -> Dict[str, float]:
    if sync_playwright is None:
        raise SkillError("playwright is required to fetch Google Finance FX rates")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(viewport={"width": 1280, "height": 900})
        try:
            eur_usd = fetch_google_finance_rate(page, "EUR-USD")
            gbp_usd = fetch_google_finance_rate(page, "GBP-USD")
            usd_cny = fetch_google_finance_rate(page, "USD-CNY")
        finally:
            browser.close()
    return {
        "USD": 1.0,
        "EUR": eur_usd,
        "GBP": gbp_usd,
        "CNY": 1.0 / usd_cny if usd_cny else 0.0,
    }


def normalize_price_to_usd(label: Optional[str], fx_rates_to_usd: Dict[str, float]) -> Optional[float]:
    values_usd: List[float] = []
    for amount, currency in extract_price_points(label):
        rate = fx_rates_to_usd.get(currency)
        if rate:
            values_usd.append(amount * rate)
    if not values_usd:
        return None
    return min(values_usd)


def normalize_handle(value: str) -> str:
    value = value.strip()
    if value.startswith("https://x.com/") or value.startswith("http://x.com/"):
        value = re.sub(r"^https?://x\.com/", "", value)
        value = value.split("?")[0].split("/")[0]
    return value.lstrip("@").strip()


def normalize_targets(raw_targets: Iterable[Any]) -> List[Target]:
    normalized: List[Target] = []
    seen = set()
    for item in raw_targets:
        if isinstance(item, str):
            handle = normalize_handle(item)
            price = None
            profile_url = f"https://x.com/{handle}"
        elif isinstance(item, dict):
            handle = normalize_handle(item.get("handle") or item.get("profile_url") or item.get("url") or "")
            price = item.get("price")
            profile_url = item.get("profile_url") or item.get("url") or f"https://x.com/{handle}"
        else:
            continue
        if not handle or handle.lower() in seen:
            continue
        seen.add(handle.lower())
        normalized.append(Target(handle=handle, profile_url=profile_url, price=price))
    return normalized


def load_targets(args: argparse.Namespace) -> List[Target]:
    raw_targets: List[Any] = []
    if args.targets_file:
        raw_targets.extend(json.loads(Path(args.targets_file).read_text()))
    if args.targets:
        raw_targets.extend([part for part in args.targets.split(",") if part.strip()])
    return normalize_targets(raw_targets)


def bootstrap_guest_session() -> Tuple[requests.Session, Dict[str, str]]:
    session = requests.Session()
    session.headers.update({"user-agent": "Mozilla/5.0"})
    html = session.get("https://x.com", timeout=30).text
    main_js_url = re.search(r'https://abs\.twimg\.com/responsive-web/client-web/main\.[^"]+\.js', html)
    if not main_js_url:
        raise SkillError("Could not find X main.js")
    main_js = session.get(main_js_url.group(0), timeout=30).text
    bearer_match = re.search(r'Bearer (AAAAA[^"\\\']+)', main_js)
    if not bearer_match:
        raise SkillError("Could not extract X bearer token")
    session.headers["authorization"] = f"Bearer {bearer_match.group(1)}"
    session.headers["x-twitter-active-user"] = "yes"
    session.headers["x-twitter-client-language"] = "en"
    guest = session.post("https://api.twitter.com/1.1/guest/activate.json", timeout=30)
    guest.raise_for_status()
    session.headers["x-guest-token"] = guest.json()["guest_token"]
    ops = {
        "bearer_token": bearer_match.group(1),
        "UserByScreenName": re.search(r'queryId:"([^"]+)",operationName:"UserByScreenName"', main_js).group(1),
        "UserTweets": re.search(r'queryId:"([^"]+)",operationName:"UserTweets"', main_js).group(1),
    }
    return session, ops


def gql_fetch(session: requests.Session, url: str, params: Dict[str, str]) -> Dict[str, Any]:
    response = session.get(url, params=params, timeout=30)
    response.raise_for_status()
    data = response.json()
    if "errors" in data:
        raise SkillError(json.dumps(data["errors"], ensure_ascii=False))
    return data


def get_user(session: requests.Session, ops: Dict[str, str], handle: str) -> Dict[str, Any]:
    return gql_fetch(
        session,
        f"https://x.com/i/api/graphql/{ops['UserByScreenName']}/UserByScreenName",
        {
            "variables": compact_json({"screen_name": handle, "withSafetyModeUserFields": True}),
            "features": compact_json(USER_BY_SCREEN_NAME_FEATURES),
            "fieldToggles": compact_json(USER_BY_SCREEN_NAME_FIELD_TOGGLES),
        },
    )["data"]["user"]["result"]


def get_timeline_page(session: requests.Session, ops: Dict[str, str], user_id: str, cursor: Optional[str]) -> Dict[str, Any]:
    variables: Dict[str, Any] = {
        "userId": user_id,
        "count": 100,
        "includePromotedContent": False,
        "withQuickPromoteEligibilityTweetFields": True,
        "withVoice": True,
        "withV2Timeline": True,
    }
    if cursor:
        variables["cursor"] = cursor
    return gql_fetch(
        session,
        f"https://x.com/i/api/graphql/{ops['UserTweets']}/UserTweets",
        {
            "variables": compact_json(variables),
            "features": compact_json(USER_TWEETS_FEATURES),
            "fieldToggles": compact_json(USER_TWEETS_FIELD_TOGGLES),
        },
    )


def extract_entries(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    instructions = data["data"]["user"]["result"]["timeline"]["timeline"]["instructions"]
    entries: List[Dict[str, Any]] = []
    for instruction in instructions:
        if "entries" in instruction:
            entries.extend(instruction["entries"])
        elif "entry" in instruction:
            entries.append(instruction["entry"])
    return entries


def extract_bottom_cursor(entries: List[Dict[str, Any]]) -> Optional[str]:
    for entry in entries:
        content = entry.get("content", {})
        if content.get("cursorType") == "Bottom":
            return content.get("value")
    return None


def unwrap_tweet_result(result: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not result:
        return None
    if result.get("__typename") == "Tweet":
        return result
    if "tweet" in result:
        return unwrap_tweet_result(result["tweet"])
    if "tweet_results" in result:
        return unwrap_tweet_result((result["tweet_results"] or {}).get("result"))
    return None


def extract_text(tweet: Dict[str, Any]) -> str:
    article_result = (((tweet.get("article") or {}).get("article_results") or {}).get("result") or {})
    blocks = (((article_result.get("content_state") or {}).get("blocks")) or [])
    article_text = "\n".join(block.get("text", "").strip() for block in blocks if block.get("text"))
    if article_text:
        return article_text.strip()
    return ((tweet.get("legacy") or {}).get("full_text") or "").strip()


def classify_post(tweet: Dict[str, Any]) -> str:
    legacy = tweet.get("legacy") or {}
    if legacy.get("retweeted_status_result"):
        return "retweet"
    if legacy.get("in_reply_to_status_id_str"):
        return "reply"
    if tweet.get("article"):
        return "article"
    if tweet.get("note_tweet"):
        return "note"
    if legacy.get("is_quote_status"):
        return "quote"
    return "post"


def is_thread_continuation(tweet: Dict[str, Any]) -> bool:
    legacy = tweet.get("legacy") or {}
    text = extract_text(tweet).lstrip()
    self_id = (((tweet.get("core") or {}).get("user_results") or {}).get("result") or {}).get("rest_id")
    return bool(
        legacy.get("in_reply_to_user_id_str") == self_id
        or text.startswith(tuple(f"{i}/" for i in range(2, 10)))
        or text.startswith("2.")
    )


def normalize_candidate(tweet: Dict[str, Any], handle: str, source_used: str) -> Dict[str, Any]:
    legacy = tweet["legacy"]
    views = int(((tweet.get("views") or {}).get("count")) or 0)
    likes = int(legacy.get("favorite_count", 0))
    retweets = int(legacy.get("retweet_count", 0))
    replies = int(legacy.get("reply_count", 0))
    engagement = likes + retweets + replies
    return {
        "id": tweet["rest_id"],
        "handle": handle,
        "url": f"https://x.com/{handle}/status/{tweet['rest_id']}",
        "created_at": datetime.strptime(legacy["created_at"], "%a %b %d %H:%M:%S %z %Y").astimezone(timezone.utc).isoformat(),
        "text_preview": extract_text(tweet).splitlines()[0][:200] if extract_text(tweet) else "",
        "post_type": classify_post(tweet),
        "impressions": views,
        "likes": likes,
        "retweets": retweets,
        "replies": replies,
        "engagement": engagement,
        "engagement_rate": round(engagement / views, 6) if views else None,
        "source_used": source_used,
        "is_reply": legacy.get("in_reply_to_status_id_str") is not None,
        "is_retweet": bool(legacy.get("retweeted_status_result")),
        "is_thread_continuation": is_thread_continuation(tweet),
    }


def collect_guest_candidates(
    session: requests.Session,
    ops: Dict[str, str],
    handle: str,
    source_used: str = "guest_public_graphql",
    max_pages: int = 60,
) -> Tuple[Dict[str, Any], List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    profile = get_user(session, ops, handle)
    user_id = profile["rest_id"]
    candidates: List[Dict[str, Any]] = []
    selected: List[Dict[str, Any]] = []
    excluded: List[Dict[str, Any]] = []
    raw_pages: List[Dict[str, Any]] = []
    seen = set()
    cursor = None

    for page_no in range(1, max_pages + 1):
        data = get_timeline_page(session, ops, user_id, cursor)
        entries = extract_entries(data)
        raw_pages.append(
            {
                "page": page_no,
                "entry_count": len(entries),
                "entry_ids": [entry.get("entryId") for entry in entries[:15]],
            }
        )
        for entry in entries:
            item = (entry.get("content") or {}).get("itemContent") or {}
            tweet = unwrap_tweet_result((item.get("tweet_results") or {}).get("result"))
            if not tweet:
                continue
            normalized = normalize_candidate(tweet, handle, source_used)
            if normalized["id"] in seen:
                continue
            seen.add(normalized["id"])
            candidates.append(normalized)
            reason = None
            if normalized["post_type"] not in ALLOWED_POST_TYPES:
                reason = normalized["post_type"]
            elif normalized["is_reply"]:
                reason = "reply"
            elif normalized["is_retweet"]:
                reason = "retweet"
            elif normalized["is_thread_continuation"]:
                reason = "thread_continuation"
            if reason:
                record = dict(normalized)
                record["selection_status"] = "excluded"
                record["exclusion_reason"] = reason
                excluded.append(record)
            else:
                record = dict(normalized)
                record["selection_status"] = "selected"
                record["exclusion_reason"] = ""
                selected.append(record)
        cursor = extract_bottom_cursor(entries)
        if len(selected) >= 20 or not cursor:
            break
        time.sleep(1)

    selected.sort(key=lambda item: item["created_at"], reverse=True)
    return profile, candidates, selected[:20], excluded, raw_pages


def maybe_launch_browser(headed: bool, browser_profile: Optional[str]) -> Optional[Tuple[Any, Any]]:
    if sync_playwright is None:
        return None
    p = sync_playwright().start()
    user_data_dir = browser_profile or tempfile.mkdtemp(prefix="yinch-x-browser-")
    context = p.chromium.launch_persistent_context(
        user_data_dir=user_data_dir,
        headless=not headed,
        channel="chrome",
        viewport={"width": 1440, "height": 960},
    )
    return p, context


def build_authenticated_session(bearer_token: str, headed: bool, browser_profile: Optional[str]) -> Optional[requests.Session]:
    if sync_playwright is None:
        return None
    launched = maybe_launch_browser(headed=headed, browser_profile=browser_profile)
    if launched is None:
        return None
    p, context = launched
    try:
        page = context.new_page()
        page.goto("https://x.com/home", wait_until="domcontentloaded")
        page.wait_for_timeout(3000)
        if "login" in page.url:
            if not headed:
                return None
            print("[x-kol] Complete X login in the opened browser, then press Enter here.", flush=True)
            input()
            page.goto("https://x.com/home", wait_until="domcontentloaded")
            page.wait_for_timeout(3000)
            if "login" in page.url:
                return None
        session = requests.Session()
        session.headers.update(
            {
                "authorization": f"Bearer {bearer_token}",
                "x-twitter-active-user": "yes",
                "x-twitter-client-language": "en",
                "user-agent": "Mozilla/5.0",
                "referer": "https://x.com/",
            }
        )
        for cookie in context.cookies("https://x.com"):
            session.cookies.set(cookie["name"], cookie["value"], domain=cookie.get("domain"), path=cookie.get("path"))
            if cookie["name"] == "ct0":
                session.headers["x-csrf-token"] = cookie["value"]
                session.headers["x-twitter-auth-type"] = "OAuth2Session"
        if "x-csrf-token" not in session.headers:
            return None
        return session
    finally:
        context.close()
        p.stop()


def collect_authenticated_candidates(
    session: requests.Session,
    ops: Dict[str, str],
    handle: str,
    max_pages: int = 60,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    profile = get_user(session, ops, handle)
    user_id = profile["rest_id"]
    candidates: List[Dict[str, Any]] = []
    raw_pages: List[Dict[str, Any]] = []
    seen = set()
    cursor = None
    for page_no in range(1, max_pages + 1):
        data = get_timeline_page(session, ops, user_id, cursor)
        entries = extract_entries(data)
        raw_pages.append(
            {
                "page": page_no,
                "entry_count": len(entries),
                "entry_ids": [entry.get("entryId") for entry in entries[:15]],
                "source": "logged_in_x_browser_graphql",
            }
        )
        for entry in entries:
            item = (entry.get("content") or {}).get("itemContent") or {}
            tweet = unwrap_tweet_result((item.get("tweet_results") or {}).get("result"))
            if not tweet:
                continue
            normalized = normalize_candidate(tweet, handle, "logged_in_x_browser_graphql")
            if normalized["id"] in seen:
                continue
            seen.add(normalized["id"])
            candidates.append(normalized)
        cursor = extract_bottom_cursor(entries)
        if not cursor:
            break
        time.sleep(1)
    candidates.sort(key=lambda item: item["created_at"], reverse=True)
    return candidates, raw_pages


def merge_candidate_sets(
    base_candidates: List[Dict[str, Any]],
    base_selected: List[Dict[str, Any]],
    base_excluded: List[Dict[str, Any]],
    extra_candidates: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    merged_candidates = list(base_candidates)
    merged_selected = list(base_selected)
    merged_excluded = list(base_excluded)
    seen = {item["id"] for item in base_candidates}
    for normalized in extra_candidates:
        if normalized["id"] in seen:
            continue
        seen.add(normalized["id"])
        merged_candidates.append(normalized)
        reason = None
        if normalized["post_type"] not in ALLOWED_POST_TYPES:
            reason = normalized["post_type"]
        elif normalized["is_reply"]:
            reason = "reply"
        elif normalized["is_retweet"]:
            reason = "retweet"
        elif normalized["is_thread_continuation"]:
            reason = "thread_continuation"
        if reason:
            record = dict(normalized)
            record["selection_status"] = "excluded"
            record["exclusion_reason"] = reason
            merged_excluded.append(record)
        else:
            record = dict(normalized)
            record["selection_status"] = "selected"
            record["exclusion_reason"] = ""
            merged_selected.append(record)
    merged_selected.sort(key=lambda item: item["created_at"], reverse=True)
    return merged_candidates, merged_selected[:20], merged_excluded


def discover_targets(topic: str, count: int, headed: bool, browser_profile: Optional[str], discovery_dir: Path) -> List[Target]:
    queries = [f"{topic} lang:en", f"{topic} lang:zh", topic]
    write_json(discovery_dir / "search_queries.json", {"queries": queries})
    discovered: List[Dict[str, Any]] = []
    if sync_playwright is None:
        raise SkillError("Discovery mode requires playwright")

    launched = maybe_launch_browser(headed=True if headed else True, browser_profile=browser_profile)
    if launched is None:
        raise SkillError("Could not launch browser for discovery mode")
    p, context = launched
    try:
        page = context.new_page()
        handles_seen = set()
        for query in queries:
            page.goto(f"https://x.com/search?q={requests.utils.quote(query)}&src=typed_query&f=user", wait_until="domcontentloaded")
            page.wait_for_timeout(4000)
            for _ in range(8):
                hrefs = page.locator("a[href^='/']").evaluate_all("(nodes) => nodes.map((node) => node.getAttribute('href'))")
                for href in hrefs:
                    if not href or href.count("/") != 1:
                        continue
                    handle = href.strip("/").split("?")[0]
                    if not handle or handle.lower() in handles_seen:
                        continue
                    if handle.lower() in {"home", "explore", "search", "i", "messages", "notifications", "settings"}:
                        continue
                    handles_seen.add(handle.lower())
                    discovered.append({"handle": handle, "profile_url": f"https://x.com/{handle}", "query": query})
                    if len(discovered) >= count * 4:
                        break
                if len(discovered) >= count * 4:
                    break
                page.mouse.wheel(0, 3000)
                page.wait_for_timeout(1200)
            if len(discovered) >= count * 4:
                break
    finally:
        context.close()
        p.stop()

    write_json(discovery_dir / "candidate_accounts.json", {"topic": topic, "candidates": discovered})
    return normalize_targets(discovered)


def summarize_posts(
    posts: List[Dict[str, Any]],
    followers: int,
    price: Optional[str],
    product_context: str,
    bio: str,
    source_used: str,
    fx_rates_to_usd: Optional[Dict[str, float]] = None,
) -> Dict[str, Any]:
    impressions = [item["impressions"] for item in posts]
    engagements = [item["engagement"] for item in posts]
    rates = [item["engagement_rate"] for item in posts if item["engagement_rate"] is not None]
    average_impression = round(sum(impressions) / len(impressions))
    median_impression = round(statistics.median(impressions))
    average_engagement = round(sum(engagements) / len(engagements))
    median_engagement = round(statistics.median(engagements))
    engagement_rate = round(sum(rates) / len(rates), 4) if rates else 0.0
    median_engagement_rate = round(statistics.median(rates), 4) if rates else 0.0
    price_value, price_currency = parse_price(price)
    cost_per_impression = round(price_value / average_impression, 6) if price_value and average_impression else None
    price_usd = normalize_price_to_usd(price, fx_rates_to_usd or {}) if price else None
    usd_cpm = round(price_usd / (average_impression / 1000), 2) if price_usd and average_impression else None
    fit_score = "Medium"
    combined_text = " ".join(item["text_preview"].lower() for item in posts)
    ctx = product_context.lower()
    bio_text = bio.lower()
    if ctx and (ctx in combined_text or ctx in bio_text):
        fit_score = "High"
    elif engagement_rate < 0.005:
        fit_score = "Low"
    assessment = f"Strict article/post sample={len(posts)}; source={source_used}; fit={fit_score}; avg impression={average_impression}; engagement rate={engagement_rate:.2%}."
    return {
        "followers": followers,
        "price": price or "",
        "source_used": source_used,
        "sample_size": len(posts),
        "average_impression": average_impression,
        "median_impression": median_impression,
        "average_engagement": average_engagement,
        "median_engagement": median_engagement,
        "engagement_rate": engagement_rate,
        "median_engagement_rate": round(median_engagement_rate, 4),
        "cost_per_impression": f"{price_currency}{cost_per_impression:.6f}" if cost_per_impression is not None else "",
        "usd_cpm": f"{usd_cpm:.2f}" if usd_cpm is not None else "",
        "assessment": assessment,
    }


def export_workbook(summary_rows: List[Dict[str, Any]], detail_rows: List[Dict[str, Any]], output_path: Path) -> None:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    detail_ws = wb.create_sheet("Details")
    summary_ws.append(SUMMARY_HEADERS)
    detail_ws.append(DETAIL_HEADERS)
    for ws in (summary_ws, detail_ws):
        for cell in ws[1]:
            cell.font = Font(bold=True)
    for row in summary_rows:
        summary_ws.append([row.get(header, "") for header in SUMMARY_HEADERS])
    for row in detail_rows:
        detail_ws.append([row.get(header, "") for header in DETAIL_HEADERS])
    ensure_dir(output_path.parent)
    wb.save(output_path)


def build_run_tree(root: Path) -> Dict[str, Path]:
    return {
        "root": ensure_dir(root),
        "input": ensure_dir(root / "input"),
        "environment": ensure_dir(root / "environment"),
        "discovery": ensure_dir(root / "discovery"),
        "collection": ensure_dir(root / "collection"),
        "analysis": ensure_dir(root / "analysis"),
        "deliverables": ensure_dir(root / "deliverables"),
    }


def main() -> None:
    args = parse_args()
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    cwd = Path.cwd()
    run_root = cwd / "yinch-auto-mkt-output" / "x-kol" / timestamp
    paths = build_run_tree(run_root)

    request_payload = {
        "mode": args.mode,
        "targets": args.targets,
        "targets_file": args.targets_file,
        "topic": args.topic,
        "count": args.count,
        "product_context": args.product_context,
        "strict_post_types": sorted(ALLOWED_POST_TYPES),
        "allow_partial": args.allow_partial,
    }
    write_json(paths["input"] / "request.json", request_payload)

    dependency_report = {}
    if args.dependency_bootstrap and Path(args.dependency_bootstrap).exists():
        dependency_report = json.loads(Path(args.dependency_bootstrap).read_text())
    write_json(paths["environment"] / "dependency_report.json", dependency_report)
    write_json(
        paths["environment"] / "run_context.json",
        {
            "cwd": str(cwd),
            "platform": platform.platform(),
            "python": sys.version,
            "timestamp": timestamp,
        },
    )
    write_json(paths["discovery"] / "search_queries.json", {"queries": []})
    write_json(paths["discovery"] / "candidate_accounts.json", {"candidates": []})

    session, ops = bootstrap_guest_session()
    authenticated_session = build_authenticated_session(
        bearer_token=ops["bearer_token"],
        headed=args.headed,
        browser_profile=args.browser_profile,
    )

    if args.mode == "account_list":
        targets = load_targets(args)
        if not targets:
            raise SkillError("account_list mode requires --targets or --targets-file")
        write_json(paths["input"] / "targets.json", [target.__dict__ for target in targets])
    else:
        if not args.topic:
            raise SkillError("discovery mode requires --topic")
        targets = discover_targets(args.topic, args.count, args.headed, args.browser_profile, paths["discovery"])
        write_json(paths["input"] / "targets.json", [target.__dict__ for target in targets])

    summary_rows: List[Dict[str, Any]] = []
    detail_rows: List[Dict[str, Any]] = []
    omitted: List[Dict[str, Any]] = []
    included = 0
    fx_rates_to_usd: Dict[str, float] = {}
    if any(target.price for target in targets):
        fx_rates_to_usd = fetch_fx_rates_to_usd()
        run_context_path = paths["environment"] / "run_context.json"
        run_context = json.loads(run_context_path.read_text())
        run_context["fx_source"] = "Google Finance"
        run_context["fx_rates_to_usd"] = {key: round(value, 6) for key, value in fx_rates_to_usd.items()}
        write_json(run_context_path, run_context)

    for target in targets:
        if args.mode == "discovery" and included >= args.count:
            break
        handle_dir = ensure_dir(paths["collection"] / slugify(target.handle))
        try:
            profile, candidates, selected, excluded, raw_pages = collect_guest_candidates(session, ops, target.handle)
            if len(selected) < 20 and authenticated_session is not None:
                auth_candidates, auth_raw_pages = collect_authenticated_candidates(
                    authenticated_session,
                    ops,
                    target.handle,
                )
                candidates, selected, excluded = merge_candidate_sets(
                    candidates,
                    selected,
                    excluded,
                    auth_candidates,
                )
                raw_pages.extend(auth_raw_pages)
            profile_payload = {
                "handle": target.handle,
                "profile_url": target.profile_url,
                "display_name": profile["core"]["name"],
                "followers": profile["legacy"].get("followers_count", 0),
                "bio": profile["legacy"].get("description", ""),
            }
            write_json(handle_dir / "profile.json", profile_payload)
            write_json(handle_dir / "raw_pages.json", raw_pages)
            write_json(handle_dir / "candidates.json", candidates)
            write_json(handle_dir / "selected_posts.json", selected)
            write_json(handle_dir / "excluded_posts.json", excluded)

            if len(selected) < 20 and not args.allow_partial:
                omitted.append({"handle": target.handle, "reason": "insufficient_sample", "sample_size": len(selected)})
                continue

            summary = summarize_posts(
                selected,
                followers=profile_payload["followers"],
                price=target.price,
                product_context=args.product_context,
                bio=profile_payload["bio"],
                source_used="+".join(sorted({row["source_used"] for row in selected})) if selected else "guest_public_graphql",
                fx_rates_to_usd=fx_rates_to_usd,
            )
            summary_row = {
                "handle": target.handle,
                "display_name": profile_payload["display_name"],
                "profile_url": target.profile_url,
                **summary,
            }
            summary_rows.append(summary_row)
            included += 1

            for row in selected:
                detail_rows.append(
                    {
                        "handle": target.handle,
                        "post_id": row["id"],
                        "post_url": row["url"],
                        "created_at": row["created_at"],
                        "post_type": row["post_type"],
                        "impressions": row["impressions"],
                        "likes": row["likes"],
                        "retweets": row["retweets"],
                        "replies": row["replies"],
                        "engagement": row["engagement"],
                        "engagement_rate": row["engagement_rate"],
                        "selection_status": "selected",
                        "exclusion_reason": "",
                        "text_preview": row["text_preview"],
                    }
                )
            for row in excluded:
                detail_rows.append(
                    {
                        "handle": target.handle,
                        "post_id": row["id"],
                        "post_url": row["url"],
                        "created_at": row["created_at"],
                        "post_type": row["post_type"],
                        "impressions": row["impressions"],
                        "likes": row["likes"],
                        "retweets": row["retweets"],
                        "replies": row["replies"],
                        "engagement": row["engagement"],
                        "engagement_rate": row["engagement_rate"],
                        "selection_status": "excluded",
                        "exclusion_reason": row["exclusion_reason"],
                        "text_preview": row["text_preview"],
                    }
                )
        except Exception as exc:
            omitted.append({"handle": target.handle, "reason": str(exc)})

    workbook_payload = {"summary_rows": summary_rows, "detail_rows": detail_rows, "omitted": omitted}
    write_json(paths["analysis"] / "summary_metrics.json", {"rows": summary_rows, "omitted": omitted})
    write_json(paths["analysis"] / "workbook_payload.json", workbook_payload)
    export_workbook(summary_rows, detail_rows, paths["deliverables"] / "x_kol_analysis.xlsx")

    print(json.dumps(
        {
            "deliverable": str(paths["deliverables"] / "x_kol_analysis.xlsx"),
            "summary": str(paths["analysis"] / "summary_metrics.json"),
            "included_kols": len(summary_rows),
            "omitted_kols": omitted,
        },
        ensure_ascii=False,
    ))


if __name__ == "__main__":
    main()
